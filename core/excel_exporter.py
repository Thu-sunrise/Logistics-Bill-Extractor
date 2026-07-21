import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    "File name", "Carrier", "Booking", "Bill no.", "Shipper", "Consignee", 
    "Notify party 1", "Notify party 2", "Notify party 3", "Remark",
    "POR", "POL", "Vessel name", "Voyage number", "POD", "Place of delivery",
    "Cont no.", "Seal no.", "Total carton", "Cont type",
    "Total GW", "Tare", "Total CBM", "ATD", "Freight", "Description"
]

def export_to_excel(all_data, output_path):
    """
    Exports all extracted data to an Excel file.
    Handles column renaming, formatting, and auto-merging cells with identical data 
    (e.g., Booking number, Vessel name) across multiple rows belonging to the same PDF file.
    """
    df = pd.DataFrame(all_data, columns=COLUMNS)
    
    # Map internal names to user requested names
    rename_map = {
        "Booking": "Booking number",
        "Bill no.": "Bill number",
        "Cont no.": "container number",
        "Seal no.": "Seal",
        "Total carton": "Total quantity",
        "Total CBM": "Total volume",
        "Remark": "Remarks"
    }
    
    # Column width adjustment
    header_widths = {
        "File name": 25, "Carrier": 10, "Shipper": 35, "Consignee": 35,
        "Booking": 18, "Bill number": 18, "Notify party 1": 35,
        "Notify party 2": 35, "Notify party 3": 35,
        "Vessel name": 25, "Voyage number": 15, "POR": 20, "POL": 20, "POD": 20,
        "container number": 18, "Seal": 18, "Total quantity": 15,
        "Cont type": 15, "Total GW": 15, "Tare": 15, "Total volume": 15, "ATD": 15,
        "Freight": 25, "Description": 45, "Remarks": 45
    }

    # Reorder columns as requested (plus File name and Carrier at the start)
    export_order = [
        "File name", "Carrier", "Shipper", "Consignee", "Booking", "Bill no.", 
        "Notify party 1", "Notify party 2", "Notify party 3",
        "Vessel name", "Voyage number", "POR", "POL", "POD", 
        "Cont no.", "Seal no.", "Total carton", "Cont type", "Total GW", "Tare",
        "Total CBM", "ATD", "Freight", "Description", "Remark"
    ]
    
    df = df[export_order]
    df.rename(columns=rename_map, inplace=True)

    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='ONE Bills')

    wb = writer.book
    ws = writer.sheets['ONE Bills']

    # Header styling
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # Set top-left alignment for all cells for better readability in merged cells
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)


    # List of columns that should NOT be merged (container-specific columns)
    container_cols = ["container number", "Seal", "Total quantity", "Cont type", "Total GW", "Tare", "Total volume"]
    merge_col_indices = [list(df.columns).index(col) + 1 for col in df.columns if col not in container_cols]
    
    start_row = 2
    for i in range(1, len(df)):
        if df.iloc[i]["File name"] != df.iloc[i-1]["File name"]:
            end_row = i + 1
            if end_row > start_row:
                for col_idx in merge_col_indices:
                    ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            start_row = i + 2
            
    # Merge the final group
    end_row = len(df) + 1
    if end_row > start_row:
        for col_idx in merge_col_indices:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

    # Auto column width (max 50)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)

    writer.close()
